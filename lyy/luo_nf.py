#!/usr/bin/env python
# coding: utf-8

# In[1]:
import pandas as pd
import openpyxl as opx
from openpyxl import Workbook
from openpyxl.styles import Alignment,Font,Side,Border,NamedStyle
from openpyxl.utils import get_column_letter
import xlrd
import os
import time
import socket
from datetime import datetime, timedelta
# import logging
# logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(lineno)d  - %(message)s')
# logger = logging.getLogger(__name__)

class sht_merge:
    def __init__(self,filename):
        pth=os.getcwd()      
        self.filename=os.path.join(pth,filename)
        
    
    def rd(self):
        print("* 正在读取\"{}\" \n".format(self.filename))
        sht_0=pd.read_excel(self.filename,sheet_name=0)  #一审及执行情况
        sht_1=pd.read_excel(self.filename,sheet_name=1,usecols=[0,1,2,3,4,5]) #二审情况
        sht_2=self.slct_ValidAgent() #先整理出有效的律所列表  代理方式
        sht_3=self.slct_dyw() #先整理出有效的不动产列表    抵押物情况
        sht_4=pd.read_excel(self.filename,sheet_name=4)  #代偿金额
        
        print("完成 \n \n* 正在处理数据... \n")

        res_1=sht_0.sort_values(by="案件名称") #表一按名称排序
        cls_2=sht_1.loc[sht_1['审级']==2]  #二审
        cls_3=sht_1.loc[sht_1['审级']==3]  #再审
        cls_4=sht_1.loc[sht_1['审级']==4]  #执行

        
        res=pd.merge(res_1,cls_2,left_on='案号',right_on='一审案号',how="left",suffixes=["","_二审"]) \
                .drop(columns=["一审案号","案件名称_二审","审级"])
        res=pd.merge(res,cls_3,left_on='案号',right_on='一审案号',how="left",suffixes=["","_再审"]) \
                .drop(columns=["一审案号","案件名称_再审","审级","主办法官_再审","管辖法院_再审"])  
        res=pd.merge(res,sht_2,left_on='案号',right_on='一审案号',how="left") \
                .drop(columns=["一审案件名称"])
        res=pd.merge(res,sht_3,left_on='案件名称',right_on='案件名称',how="left")
        res=pd.merge(res,sht_4,left_on='案件名称',right_on='案件名称',how="left")

        order=["案件名称","起诉金额（元）","财务代偿余额（元）","抵押物情况","主办","副办","诉讼状态" \
                 ,"代理方式","律所名称","辅助机构"  \
                 ,"案号","案号_二审","案号_再审","执行案号","起诉日期" \
                 ,"立案日期","一审管辖法院","一审主办法官" \
                 ,"管辖法院","主办法官" \
                 ,"执行法院","执行主办法官" \
                 ,"备注"]
        
        res=res[order]
        res=res.rename(columns={"案号":"一审案号","案号_二审":"二审案号","案号_再审":"再审案号" \
                 ,"审理情况":"二审审理情况","管辖情况":"二审管辖情况","主办法官":"二审主办法官"  \
                 ,"管辖法院":"二审管辖法院"})

        # print(res)
        self.res=res
        print("完成 \n \n ")
    
    #处理代理机构表，挑选出有效的律所
    def slct_ValidAgent(self):

        sht_2=pd.read_excel(self.filename,sheet_name=2)

        dl_1=sht_2.loc[sht_2["一审代理律所名称"]==sht_2["目前有效"]] \
                        .drop(columns=["二审代理律所名称","再审代理律所名称","执行代理律所名称"]) \
                                        .rename(columns={"一审代理律所名称":"律所名称"})

        dl_2=sht_2.loc[sht_2["二审代理律所名称"]==sht_2["目前有效"]] \
                        .drop(columns=["一审代理律所名称","再审代理律所名称","执行代理律所名称"]) \
                                        .rename(columns={"二审代理律所名称":"律所名称"})

        dl_3=sht_2.loc[sht_2["再审代理律所名称"]==sht_2["目前有效"]] \
                        .drop(columns=["一审代理律所名称","二审代理律所名称","执行代理律所名称"]) \
                                        .rename(columns={"再审代理律所名称":"律所名称"})

        dl_4=sht_2.loc[sht_2["执行代理律所名称"]==sht_2["目前有效"]] \
                        .drop(columns=["一审代理律所名称","二审代理律所名称","再审代理律所名称"]) \
                                        .rename(columns={"执行代理律所名称":"律所名称"})

        dl=pd.concat([dl_1,dl_2,dl_3,dl_4],ignore_index=True).drop(columns=["目前有效"])

        dl=dl.drop_duplicates(subset="一审案号",keep="first",inplace=False)

        # print(dl)
        return dl    
    
    #不动产，按案件名称整理
    def slct_dyw(self):
        sht_3=pd.read_excel(self.filename,sheet_name=3)

        lst_case=[]
        for i in sht_3["案件名称"]:
            if i not in lst_case:
                lst_case.append(i)

        dyw=[]
        for i in lst_case:
            _dyw=""
            for txt_dyw in sht_3.loc[sht_3["案件名称"]==i]["抵押不动产位置"]:
                _dyw=_dyw+str(txt_dyw)+"；"
            _dyw=_dyw[0:-1]
            dyw.append([i,_dyw])

        lst_dyw = pd.DataFrame(dyw)
        lst_dyw.columns=["案件名称","抵押物情况"]
        lst_dyw=lst_dyw[~lst_dyw["抵押物情况"].isin(["nan"])] #反选，抵押物情况不为nan，即有抵押物情况存在的案件
        
        return lst_dyw
    
    def wt(self):
        print("* 正在生成临时数据... \n")
        wt=pd.ExcelWriter(os.path.join(os.getcwd(),'output.xlsx'))
        self.res.to_excel(wt,index=False)
        wt.save()
        # print("完成 \n \n")

class wt_excel:
    def __init__(self,filename,outputname="报表.xlsx"):
        pth=os.getcwd()        
        self.filename=os.path.join(pth,filename)
        self.outputname=os.path.join(pth,outputname)
        self.origin=os.path.join(pth,'南宁市南方融资担保有限公司诉讼案件子表.xlsm')
    
    def wt(self):
        print("* 正在写入数据，并调整格式... \n")
        wb = opx.load_workbook(self.filename)
        sht = wb['Sheet1']
        sht.insert_rows(1,2)
        sht.insert_cols(1,2)
        mrows=sht.max_row
        mcols=sht.max_column
    
        e="C"+str(mrows)
        sht['A3']='户数'
        sht['B3']='案件数'

        
        n=4 #表格数据从第4行开始
        lst=[]
        lst_grp=[]
        for rows in sht["C4":e]:
            for cells in rows:
                lst_grp.append([cells.value,n])
                if cells.value not in lst:
                    lst.append(cells.value)                
                n+=1 
                
        g=[]               
        for i in lst:
            g_0=[] 
            for j in lst_grp:
                if j[0]==i:
                    g_0.append(j[1])
            g.append(g_0)
            
        gp=[]
        col_to_merge=["A","C","E","F"] #需要合并的列坐标
        for i in g:
            if i[0]!=i[-1]:
                for j in col_to_merge:
                    gp.append(j+str(i[0])+":"+j+str(i[-1])) 
          
        for i in gp:
            sht.merge_cells(i)
        
        # 按“案件名称”填写A列户数
        n=1
        for _case in lst:
            for row in range(4,mrows+1):
                for col in range(1,4):
                    if sht["C"+str(row)].value==_case:
                        sht["A"+str(row)]=n
            n+=1

        # 按“案件数”填写B列案件数
        n=1
        for row in sht["B4:B"+str(mrows)]:
            row[0].value=n
            n+=1


        sht.merge_cells("M2:T2")
        sht.merge_cells("U2:V2")
        sht.merge_cells("W2:X2")
        sht.merge_cells("A1:Y1")    
        sht["M2"].value="一审"
        sht["U2"].value="二审"
        sht["W2"].value="执行"
        sht["A1"].value="南宁市南方融资担保有限公司诉讼案件总表" 
                    

        #调整格式     
        title_A_L=sht["A3:L3"]  
        title_Y=sht["Y3"]     

        cols_to_merge=["A","B","C","D","E","F","G","H","I","J","K","L"]
        for i in cols_to_merge:
            sht.merge_cells(i+"2:"+i+"3")

        sht.merge_cells("Y2:Y3")

        n=0
        for i in sht["A2:L2"][0]:
            i.value=title_A_L[0][n].value
            # print(i.value)
            n+=1

        sht["Y2"].value=title_Y.value



        #打包样式
        line_t = Side(style='thin', color='000000')  # 细边框
        line_m = Side(style='medium', color='000000')  # 粗边框
        num_fmt="#,##0.00"
        border0 = Border(top=line_m, bottom=line_m, left=line_m, right=line_m)
        border1 = Border(top=line_t, bottom=line_t, left=line_t, right=line_t)        
        align = Alignment(horizontal='left',vertical='center',wrap_text=True)
        align_num = Alignment(horizontal='right',vertical='center',wrap_text=True)
        align_title=Alignment(horizontal='center',vertical='center',wrap_text=True)
        sty1 = NamedStyle(name='sty1', border=border1, alignment=align)
        sty_title= NamedStyle(name='sty_title',border=border0,alignment=align_title,font=Font(bold=True,size=11))
        sty_big_title= NamedStyle(name='sty_big_title',alignment=align_title,font=Font(bold=True,size=22))
        sty_num=NamedStyle(name='sty_num',border=border1, alignment=align_num,number_format=num_fmt)
        
        sht["A1"].style=sty_big_title
        sht["M2"].style=sty_title
        sht["U2"].style=sty_title
        sht["W2"].style=sty_title
        
        #标题格式
        for r in range(2,4):
            for c in range(1,mcols+1):
                sht.cell(r,c).style=sty_title
        #正文格式
        for r in range(4,mrows+1):
            for c in range(1,mcols+1):
                sht.cell(r,c).style=sty1    
                
        #D、E列格式（数字）
        for r in range(4,mrows+1):
            for c in range(4,6):
                sht.cell(r,c).style=sty_num    
        
        #调整列宽
        #获取每一列的内容的最大宽度
        i = 0
        col_width=[]
        for col in sht.columns:
            for j in range(len(col)):
                if j == 0:
                    col_width.append(len(str(col[j].value)))
                else:
                    # 获得每列中的内容的最大宽度
                    if col_width[i] < len(str(col[j].value)):
                        col_width[i] = len(str(col[j].value))
            i = i + 1

         #设置列宽
        for i in range(len(col_width)):
             # 根据列的数字返回字母
            col_letter = get_column_letter(i+1)
             # 当宽度大于100，宽度设置为100
            if col_width[i] > 100:
                sht.column_dimensions[col_letter].width = 100
             # 只有当宽度大于10，才设置列宽
            elif col_width[i] > 15:
                sht.column_dimensions[col_letter].width = col_width[i] + 2
        
        #调整行高 
        sht.row_dimensions[1].height = 44
        sht.row_dimensions[2].height = 33
        sht.row_dimensions[3].height = 36
        sht.column_dimensions["A"].width = 6
        sht.column_dimensions["D"].width = 15
        sht.column_dimensions["E"].width = 15
        
        print("完成 \n \n * 正在生成文件：{}... \n".format(self.outputname))
            
        wb.save(self.outputname)
        print("完成报表\n")
        # a=input("按回车退出")

    def to_txt(self,today):    
        print('*正在处理数据...\n') 
        # today='2019-7-13'
        s_left=today.find('-')
        s_right=today.rfind('-')
        month=today[0:s_right] #月份
        m=today[s_left+1:s_right]
        if m[0]=='0':
            m=m[-1]   # {4}
        d=today[s_right+1:]
        if d[0]=='0':
            d=d[-1]

        day_txt=m+'月'+d+'日'  #日期 {0}
        
        df1=pd.read_excel(self.filename)
        df1['起诉日期']=pd.to_datetime(df1['起诉日期'],format="%Y-%m-%d")
        
        col_names=["户数","案件数","案件名称","起诉金额（元）","财务代偿余额（元）","抵押物情况","主办","副办","诉讼状态","代理方式","律所名称","辅助机构","一审案号","二审案号","再审案号","执行案号","起诉日期","立案日期","一审管辖法院","一审主办法官","二审管辖法院","二审主办法官","执行法院","执行主办法官","备注"]
        # df2=pd.read_excel(self.outputname,skiprows=3,names=col_names)
        
        df_before_today=df1.loc[df1['起诉日期']<=today] #累计到录入时间点        
        df1_qisuIndex=df1.set_index('起诉日期') #起诉日期索引
        df1_qisuIndex=df1_qisuIndex[month]

        #截至1月31日，包括政策性担保项目在内，我公司累计诉讼69户96件，起诉金额55034.87万元。【累计】
        hushu_total=len(list(set(df_before_today['案件名称'].values)))                        #去重后，累计户数 {1}
        num_totalcases=df_before_today.shape[0] #行数 ，即累计案件数 {2} {18}
        amt_qisu=str((df_before_today['起诉金额（元）'].sum()/10000).round(2))+"万元"  #累计起诉金额  {3}

        # 1月新增案件1起（诉都安润美泉农业科技），结清案件0起。【录入月】
        num_case_new=df1_qisuIndex.shape[0] # 录入月份的新增案件数  {5}
                                        #录入月份新增案件名称  {6}
        _case=df1_qisuIndex['案件名称'].values
        _case=list(set(_case)) #利用set函数去重
        case=''
        if _case:
            for txt in _case:
                case=case+txt+'、'
            casenames='（'+case[0:-1]+'）'     
        else:
            casenames=''
                                                         
        num_case_done=df_before_today.loc[df_before_today['诉讼状态']=='结清'].shape[0] # 录入月份的结清案件数  {7}

        #目前待开庭案件8件，涉及起诉金额4065.47万元......【累计】
        
        num_toopen_total=df_before_today.loc[df_before_today['诉讼状态']=='待开庭'].shape[0] #累计待开庭案件数   {8}
        amt_toopen_total=str((df_before_today.loc[df_before_today['诉讼状态']=='待开庭']['起诉金额（元）'].sum()/10000).round(2))+'万元' #累计待开庭案件起诉金额 {9}
        num_tojdg_total=df_before_today.loc[df_before_today['诉讼状态']=='待判决'].shape[0] #累计待判决案件数  {10}
        amt_tojdg_total=str((df_before_today.loc[df_before_today['诉讼状态']=='待判决']['起诉金额（元）'].sum()/10000).round(2))+'万元' #累计待判决案件起诉金额 {11}
        num_toexec_total=df_before_today.loc[df_before_today['诉讼状态']=='申请执行'].shape[0] #累计 准备申请执行 案件数  {12}
        amt_toexec_total=str((df_before_today.loc[df_before_today['诉讼状态']=='申请执行']['起诉金额（元）'].sum()/10000).round(2))+'万元' #累计 准备申请执行 案件起诉金额 {13}
        num_todo_total=df_before_today.loc[df_before_today['诉讼状态']=='执行'].shape[0] #累计 执行阶段 案件数 {14}
        amt_todo_total=str((df_before_today.loc[df_before_today['诉讼状态']=='执行']['起诉金额（元）'].sum()/10000).round(2))+'万元' #累计 执行阶段 案件起诉金额 {15}
        num_2nd_cases=df_before_today.loc[df_before_today['诉讼状态']=='二审'].shape[0] #二审案件数 {16}
        amt_2nd_cases=str((df_before_today.loc[df_before_today['诉讼状态']=='二审']['起诉金额（元）'].sum()/10000).round(2))+'万元' #累计 二审 案件起诉金额 {17}
        hushu_lawyerOffice=len(list(set(df_before_today.loc[df_before_today['代理方式']=='律所代理']['案件名称'].values))) #律所代理户数 {19}
        num_lawyerOffice=df_before_today.loc[df_before_today['代理方式']=='律所代理'].shape[0] #律所代理数 {20}
        amt_lawyerOffice=str((df_before_today.loc[df_before_today['代理方式']=='自己代理']['起诉金额（元）'].sum()/10000).round(2))+'万元' #累计 律所代理 案件起诉金额 {21}
        hushu_self=len(list(set(df_before_today.loc[df_before_today['代理方式']=='自己代理']['案件名称'].values))) #自行代理户数 {22}
        num_self=df_before_today.loc[df_before_today['代理方式']=='自己代理'].shape[0] #自行代理数 {23}
        amt_self=str((df_before_today.loc[df_before_today['代理方式']=='自己代理']['起诉金额（元）'].sum()/10000).round(2))+'万元' #累计 自行代理 案件起诉金额 {24}


        # {0} : day_txt
        # {1} : hushu_total
        # {2} : num_totalcases
        # {3} : amt_qisu
        # {4} : m
        # {5} : num_case_new
        # {6} : casenames
        # {7} : num_case_done
        # {8} : num_toopen_total
        # {9} : amt_toopen_total
        # {10} : num_tojdg_total
        # {11} : amt_tojdg_total
        # {12} : num_toexec_total
        # {13} : amt_toexec_total
        # {14} : num_todo_total
        # {15} : amt_todo_total
        # {16} : num_2nd_cases
        # {17} : amt_2nd_cases
        # {18} : num_totalcases
        # {19} : hushu_lawyerOffice
        # {20} : num_lawyerOffice
        # {21} : amt_lawyerOffice
        # {22} : hushu_self
        # {23} : num_self
        # {24} : amt_self

        t='''
 （三）代偿债权处置
 1.诉讼工作
 截至{0}，包括政策性担保项目在内，我公司累计诉讼{1}户{2}件，起诉金额{3}。{4}月新增案件{5}起{6}，结清案件{7}起。目前待开庭案件{8}件，涉及起诉金额{9}；待判决案件{10}件，涉及起诉金额{11}；准备申请执行案件{12}件，涉及起诉金额{13}；执行阶段案件{14}件，涉及起诉金额{15}；二审阶段案件{16}件，涉及起诉金额{17}。在{18}件诉讼案件中，采用律所风险代理方式处理的案件有{19}户{20}件，金额{21}；自行代理方式的案件{22}户{23}件，金额{24}。
                
                '''  

        txt_out=t.format(day_txt,hushu_total,num_totalcases,amt_qisu,m,num_case_new,casenames,num_case_done,num_toopen_total,amt_toopen_total, \
                            num_tojdg_total,amt_tojdg_total,num_toexec_total,amt_toexec_total,num_todo_total,amt_todo_total,num_2nd_cases,amt_2nd_cases, \
                            num_totalcases,hushu_lawyerOffice,num_lawyerOffice,amt_lawyerOffice,hushu_self,num_self,amt_self)        
        
        fn_txt=os.path.join(os.getcwd() ,month+'总结.txt')
        with open(fn_txt,'w') as f:
            f.write(txt_out)

        # input('已完成，按回车退出。')

def main():
    slct=input('请输入：\n1——生成报表\n2——生成本月总结（片段）\n')
    if slct=='1':
        luo=sht_merge("南宁市南方融资担保有限公司诉讼案件子表.xlsm")
        luo.rd()
        luo.wt()
        
        wt=wt_excel("output.xlsx")
        wt.wt()
        input('已完成报表，按回车退出。')
    elif slct=='2':
        today=input('请输入日期【YYYY-M-D】:\n')
        luo=sht_merge("南宁市南方融资担保有限公司诉讼案件子表.xlsm")
        luo.rd()
        luo.wt()
        
        wt=wt_excel("output.xlsx")
        wt.wt()
        wt.to_txt(today)
        input('已完成文字生成，按回车退出。')
    else:
        input('不在选择之内，按回车退出重新选择。')

                             
if __name__=="__main__":
    main()
