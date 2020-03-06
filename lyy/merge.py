import os
import openpyxl

class merge:
    def __init__(self,fn,sht,c1,*c):
        self.pth=os.getcwd()
        self.fn=os.path.join(self.pth,fn)
        self.inputname=fn
        self.sht=sht
        self.c1=c1
        self.c1_asc=ord(c1.lower())-96       

        if isinstance(c[0],list):  
            self._c=[]
            for i  in c[0]:
                self._c.append(i)    #如果C是通过input来输入的，因为输入的是一个List, 要经过此拆分list[0]来重新构造self._c
        else:
            self._c=c   #如是直接在类实例化时输入c的，self._c可直接被赋值为c

    def mg(self):
        wb=openpyxl.load_workbook(self.fn)
        sht=wb[self.sht]
        mrows=sht.max_row

        casename=[]
        amt=[]
        for r in range(1,mrows+1):
            for c in range(self.c1_asc,self.c1_asc+1):
                v=sht.cell(r,c).value
                amt.append([v,r])
                if v not in casename:
                    casename.append(v)

        p=[]
        for i in casename:
            q=[]
            for j in amt:
                if j[0]==i:
                    q.append(j[1])
            p.append(q)

        t=[self.c1]
        for i in self._c:        
            if i!='':
                t.append(i)
        
                
        g=[]
        for k in t:
            for i in p:
                if i[0]!=i[-1]:
                    g.append(k+str(i[0])+":"+k+str(i[-1]))


        for i in g:
            sht.merge_cells(i)

        print('完成以下合并：\n{}\n'.format(g))

        out=os.path.join(self.pth,self.inputname[0:-5]+'合并单元格后.xlsx')
        wb.save(out)

        input('合并完成。生成文件：{}\n按回车退出'.format(out))
        
        
        
def main():
    fn=input('输入文件名：')
    sht=input('输入表名：')
    col_1=input('输入第一列字母：')

    # fn='test.xlsx'
    # sht='Sheet2'
    # col_1='a'
    # col_2='b'

    cols=[]
    while True:
        _cols=input('输入其它列字母，要结束输入，请直接按回车：')
        if _cols!='':
            cols.append(_cols)
        else:
            break

    s=merge(fn,sht,col_1,cols)
    # s=merge(fn,sht,col_1,col_2,'d','k','l')

    s.mg()

if __name__=='__main__':
    # s=merge('test.xlsx','表1','a','b')
    # s.mg()
    main()